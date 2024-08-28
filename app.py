from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Necessary for flashing messages

# Configure Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'flala22@beder.edu.al'
app.config['MAIL_PASSWORD'] = 'Bederuni1234'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False

mail = Mail(app)
s = URLSafeTimedSerializer(app.secret_key)

# Load or create an Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Users"
ws.append(["Name", "Surname", "Phone Number", "Email", "Password", "Confirmed"])  # Header row with "Confirmed" status

@app.route('/')
def index():
    return render_template('signup.html')

@app.route('/signup', methods=['POST'])
def signup():
    name = request.form['name']
    surname = request.form['surname']
    phone = request.form['phone']
    email = request.form['email']
    password = request.form['password']

    # Save user data with confirmation status set to False
    ws.append([name, surname, phone, email, password, False])
    wb.save("users.xlsx")

    # Generate a token for email confirmation
    token = s.dumps(email, salt='email-confirm')

    # Send confirmation email
    msg = Message('Confirm your Email', sender='flala22@beder.edu.al', recipients=[email])
    link = url_for('confirm_email', token=token, _external=True)
    msg.body = f'We appreciate your interest in joining us. To complete your registration process, please click the link below: {link}'
    mail.send(msg)

    flash("A confirmation email has been sent to your email address. Please confirm your email to log in.")
    return redirect(url_for('signin'))

@app.route('/confirm_email/<token>')
def confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm', max_age=3600)  # Token valid for 1 hour
    except SignatureExpired:
        return '<h1>The token is expired!</h1>'

    # Mark the user's email as confirmed in the Excel file
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[3].value == email:
            row[5].value = True  # Set "Confirmed" status to True
            wb.save("users.xlsx")
            return render_template('confirmation.html', email=email)

    return '<h1>Email not found!</h1>'


@app.route('/signin')
def signin():
    return render_template('signin.html')

@app.route('/signin', methods=['POST'])
def signin_post():
    email = request.form['email']
    password = request.form['password']

    # Load the Excel file and check credentials
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == email and row[4] == password:
            if row[5]:  # Check if the email is confirmed
                return redirect(url_for('success'))
            else:
                flash("Please confirm your email before logging in.")
                return redirect(url_for('signin'))

    flash("Incorrect email or password. Please try again.")
    return redirect(url_for('signin'))

@app.route('/success')
def success():
    return "Sign in successful! Welcome!"

@app.route('/forgot-password')
def forgot_password():
    return "Password recovery process initiated. (This is a placeholder page.)"

if __name__ == '__main__':
    app.run(debug=True)
